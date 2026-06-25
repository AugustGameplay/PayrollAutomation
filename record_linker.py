#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
 VINCULADOR DE REGISTROS  ·  Record Linker  (v2 · name-first)
 Primary Staffing Inc.
----------------------------------------------------------------------------
 Enlaza dos archivos de personas usando el NOMBRE como clave principal
 (no depende del Employee ID). Estrategia:
    1) Nombre (fuzzy)                         -> clave principal
    2) + Monto de confirmacion (si coincide)  -> sube la confianza
    3) + ID secundario SSN/EmployeeID         -> solo si AMBOS archivos lo
                                                 comparten; confirma al 100%
    4) Nombres ambiguos / dudosos / sin match -> REVISAR MANUAL
 Salida: CSV consolidado con metodo, % de confianza, margen de error,
 % de similitud de nombre, si el monto coincide, si el nombre es ambiguo,
 y el estado (AUTOMATICO / REVISAR MANUAL).
============================================================================
 Requisitos (una sola vez):  pip install pandas openpyxl rapidfuzz
 Ejecutar:                    python record_linker.py
============================================================================
"""

import os
import re
import csv
import queue
import threading
import unicodedata
import datetime as dt

import pandas as pd

try:
    from rapidfuzz import fuzz, process
    HAVE_RF = True
except Exception:
    HAVE_RF = False

# ===========================================================================
#  MOTOR DE VINCULACION  (independiente de la interfaz)
# ===========================================================================

HINTS_NAME   = ["employee", "name", "nombre", "full name", "empleado"]
HINTS_AMOUNT = ["item pay", "earn amt", "gross", "net pay", "net amount",
                "total pay", "importe", "monto", "pay", "amount"]
HINTS_SID    = ["ssn", "social security", "seguro social", "employee id",
                "emp id", "ee number"]
HINTS_RATE   = ["pay rate", "rate", "tarifa", "hourly rate"]
HINTS_PERIOD = ["week worked", "week", "periodo", "period", "check date",
                "fecha", "semana", "pay period"]

AMT_TOL  = 0.05    # tolerancia de monto
RATE_TOL = 0.01    # tolerancia de tarifa


def _ratio(a, b):
    if HAVE_RF:
        return fuzz.token_sort_ratio(a, b)
    sa, sb = set(a.split()), set(b.split())
    if not sa or not sb:
        return 0
    return 100.0 * len(sa & sb) / len(sa | sb)


def norm_name(v):
    """minusculas, sin acentos, sin puntuacion, tokens ordenados."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace(",", " ")
    s = re.sub(r"[^a-z ]", " ", s)
    return " ".join(sorted(s.split()))


def sid_kind(colname):
    """Identifica el tipo de un ID secundario por el nombre de la columna."""
    cl = str(colname).lower()
    if "ssn" in cl or "social" in cl or "seguro" in cl:
        return "ssn"
    if "employee id" in cl or "emp id" in cl or "ee number" in cl or cl.strip() == "id":
        return "empid"
    return "other"


def norm_sid(v, kind):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).upper()
    s = re.sub(r"EMP\s*ID\s*:?", "", s)
    s = re.sub(r"EE\s*NUMBER\s*:?", "", s)
    if kind == "ssn":
        return re.sub(r"\D", "", s)          # solo digitos
    return re.sub(r"[^A-Z0-9]", "", s)       # alfanumerico


def norm_period(v):
    """Normaliza una fecha/semana a 'YYYY-MM-DD' (o texto limpio si no es fecha)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        return pd.to_datetime(v).strftime("%Y-%m-%d")
    except Exception:
        s = str(v).strip()
        s = re.sub(r"(check\s*dt|period|week)\s*:?", "", s, flags=re.I).strip()
        return s


def to_num(v):
    try:
        return round(float(v), 2)
    except Exception:
        return None


JUNK_RE = re.compile(
    r"^\s*(totals?\b|total for|totals for|grand total|report total|"
    r"company name|selection criteria|page \d)", re.I)


def is_person(name):
    if name is None:
        return False
    s = str(name).strip()
    if not s or JUNK_RE.match(s):
        return False
    return bool(re.search(r"[A-Za-z]", s))


def parse_blocks(df, name_col, amount_col=None, sid_col=None,
                 rate_col=None, period_col=None):
    """Convierte un reporte en UN registro por persona. Maneja DOS formatos:
      · Por bloques: una fila cabecera con el nombre + filas de detalle vacias
        (ej. Register New/Regular).
      · Plano: el nombre se repite en cada fila (ej. Invoice/Payroll de Avionte).
    Suma el monto, toma la tarifa mas frecuente y la semana/periodo de cada
    persona. Separa homonimos seguidos cuando su ID secundario difiere."""
    kind = sid_kind(sid_col) if sid_col else None

    name_ff = df[name_col].ffill().fillna("")
    if sid_col:
        sid_ff = df[sid_col].ffill().fillna("").map(lambda v: norm_sid(v, kind))
    else:
        sid_ff = pd.Series([""] * len(df), index=df.index)

    boundary = (name_ff != name_ff.shift()) | (sid_ff != sid_ff.shift())
    groups = boundary.cumsum()

    records = []
    for _, block in df.groupby(groups, sort=True):
        nm_vals = block[name_col].dropna()
        nm = (str(nm_vals.iloc[0]).strip() if len(nm_vals)
              else str(name_ff.loc[block.index[0]]).strip())
        if not is_person(nm):
            continue
        rec = {"_name": nm, "_namekey": norm_name(nm)}

        for c in df.columns:
            col = block[c].dropna()
            rec[c] = col.iloc[0] if len(col) else ""

        # monto: SUMA del bloque
        if amount_col:
            s = pd.to_numeric(block[amount_col], errors="coerce").sum()
            rec["_amount"] = round(float(s), 2)
            rec[amount_col] = round(float(s), 2)
        else:
            rec["_amount"] = None

        # tarifa: valor mas frecuente del bloque (ignora 0 / vacios)
        if rate_col:
            rt = pd.to_numeric(block[rate_col], errors="coerce").dropna()
            rt = rt[rt > 0]
            rec["_rate"] = round(float(rt.mode().iloc[0]), 2) if len(rt) else None
        else:
            rec["_rate"] = None

        # semana/periodo: primer valor del bloque (y conjunto de semanas)
        if period_col:
            pv = block[period_col].dropna()
            rec["_period"] = norm_period(pv.iloc[0]) if len(pv) else ""
            rec["_periods"] = sorted({norm_period(x) for x in pv}) if len(pv) else []
        else:
            rec["_period"], rec["_periods"] = "", []

        if sid_col:
            sv = block[sid_col].dropna()
            rec["_sid"] = norm_sid(sv.iloc[0], kind) if len(sv) else ""
            rec["_sidkind"] = kind
        else:
            rec["_sid"], rec["_sidkind"] = "", None

        records.append(rec)
    return records


def confidence(name_sim, amt, rate, period, idm, ambiguous):
    """Confianza 0-100. Base = nombre; cada confirmacion (monto, tarifa,
    semana, ID) la sube. Un homonimo sin resolver se topa en 78 -> manual."""
    if idm:
        return 100.0
    c = float(name_sim)
    if amt:
        c += 10.0
    if rate:
        c += 6.0
    if period:
        c += 4.0
    c = min(100.0, c)
    if ambiguous:
        c = min(c, 78.0)
    return round(c, 1)


def _evidence(a, b, shared_id):
    """Compara dos registros y devuelve banderas de coincidencia."""
    amt = (a["_amount"] is not None and b["_amount"] is not None
           and abs(a["_amount"] - b["_amount"]) <= AMT_TOL)
    rate = (a["_rate"] is not None and b["_rate"] is not None
            and abs(a["_rate"] - b["_rate"]) <= RATE_TOL)
    period = bool(a["_period"] and b["_period"] and a["_period"] == b["_period"])
    idm = bool(shared_id and a["_sid"] and b["_sid"] and a["_sid"] == b["_sid"])
    return amt, rate, period, idm


def link(records_a, records_b, threshold, progress_cb=None):
    """Vincula A contra B por NOMBRE y resuelve empates (homonimos) usando
    monto + tarifa + semana + ID secundario combinados."""
    ka = records_a[0]["_sidkind"] if records_a else None
    kb = records_b[0]["_sidkind"] if records_b else None
    shared_id = (ka is not None and ka == kb and ka != "other")

    b_names = [rb["_namekey"] for rb in records_b]
    matched_obj, matched_sid = set(), set()
    rows = []
    n = len(records_a)

    for i, ra in enumerate(records_a):
        method, b = "SIN MATCH", None
        name_sim, amt_m, rate_m, per_m, id_m, ambiguous = 0, False, False, False, False, False
        homonym = False

        if b_names:
            if HAVE_RF:
                cands = process.extract(ra["_namekey"], b_names,
                                        scorer=fuzz.token_sort_ratio, limit=6)
            else:
                cands = sorted(
                    [(bn, _ratio(ra["_namekey"], bn), j) for j, bn in enumerate(b_names)],
                    key=lambda x: x[1], reverse=True)[:6]
            cands = [c for c in cands if c[1] >= 80]

            scored = []
            for bn, nsim, idx in cands:
                cb = records_b[idx]
                am, rt, pr, idd = _evidence(ra, cb, shared_id)
                combined = nsim + (10 if am else 0) + (6 if rt else 0) + \
                           (4 if pr else 0) + (40 if idd else 0)
                scored.append((combined, nsim, am, rt, pr, idd, cb))
            scored.sort(key=lambda x: x[0], reverse=True)

            if scored:
                best = scored[0]
                _, name_sim, amt_m, rate_m, per_m, id_m, b = best
                # ¿cuantos candidatos comparten practicamente el mismo nombre?
                same_name = sum(1 for s in scored if s[1] >= 90)
                homonym = same_name > 1
                # ¿homonimo sin resolver? la evidencia no separa al ganador
                if len(scored) > 1:
                    sec = scored[1]
                    if (name_sim >= 90 and sec[1] >= 90
                            and (best[0] - sec[0]) < 6):
                        ambiguous = True
                if not id_m and name_sim < 80:
                    b, method = None, "SIN MATCH"

        if b is not None:
            method = ("NOMBRE+ID" if id_m else
                      "+".join(["NOMBRE"] +
                               (["MONTO"] if amt_m else []) +
                               (["TARIFA"] if rate_m else []) +
                               (["SEMANA"] if per_m else [])))

        conf = confidence(name_sim, amt_m, rate_m, per_m, id_m, ambiguous)
        status = "AUTOMATICO" if conf >= threshold else "REVISAR MANUAL"
        if b is not None and conf >= threshold:
            matched_obj.add(id(b))
            if b["_sid"]:
                matched_sid.add(b["_sid"])

        rows.append({"ra": ra, "rb": b, "method": method, "conf": conf,
                     "margin": round(100 - conf, 1), "status": status,
                     "name_sim": round(name_sim, 1), "amt_match": amt_m,
                     "rate_match": rate_m, "per_match": per_m,
                     "id_match": id_m, "ambiguous": ambiguous, "homonym": homonym})
        if progress_cb and (i % 15 == 0 or i == n - 1):
            progress_cb(i + 1, n, ra["_name"])

    b_only = [rb for rb in records_b
              if id(rb) not in matched_obj
              and not (rb["_sid"] and rb["_sid"] in matched_sid)]
    return rows, b_only, shared_id


def build_output(rows, b_only, cols_a, cols_b):
    out = []
    for r in rows:
        ra, rb = r["ra"], r["rb"]
        row = {"Origen": "A"}
        for c in cols_a:
            row[f"A_{c}"] = ra.get(c, "")
        for c in cols_b:
            row[f"B_{c}"] = (rb.get(c, "") if rb is not None else "")
        row["Metodo"] = r["method"]
        row["Confianza_%"] = r["conf"]
        row["Margen_error_%"] = r["margin"]
        row["Similitud_nombre_%"] = r["name_sim"]
        row["Monto_coincide"] = "SI" if r["amt_match"] else "NO"
        row["Tarifa_coincide"] = "SI" if r["rate_match"] else "NO"
        row["Semana_coincide"] = "SI" if r["per_match"] else "NO"
        row["ID_coincide"] = "SI" if r["id_match"] else "NO"
        row["Nombre_ambiguo"] = "SI" if r["ambiguous"] else "NO"
        row["Homonimo_resuelto"] = "SI" if (r["homonym"] and not r["ambiguous"]) else "NO"
        row["Estado"] = r["status"]
        out.append(row)
    for rb in b_only:
        row = {"Origen": "B SIN PAREJA"}
        for c in cols_a:
            row[f"A_{c}"] = ""
        for c in cols_b:
            row[f"B_{c}"] = rb.get(c, "")
        row.update({"Metodo": "SIN MATCH", "Confianza_%": 0, "Margen_error_%": 100,
                    "Similitud_nombre_%": 0, "Monto_coincide": "NO",
                    "Tarifa_coincide": "NO", "Semana_coincide": "NO", "ID_coincide": "NO",
                    "Nombre_ambiguo": "NO", "Homonimo_resuelto": "NO", "Estado": "REVISAR MANUAL"})
        out.append(row)
    return out


def load_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        sep = "\t" if ext == ".tsv" else None
        return pd.read_csv(path, sep=sep, engine="python", dtype=str)
    return pd.read_excel(path, sheet_name=0)


# --- Vista ORGANIZADA: columnas comparables lado a lado -------------------

def build_readable(rows, b_only, m, extra_b=None):
    """Arma una tabla legible donde cada campo de A queda JUNTO a su par de B.
    m = columnas elegidas: na,aa,ra,pa,sa (A) y nb,ab,rb,pb,sb (B).
    extra_b = columnas extra de B a mostrar al final (Location, Position...)."""
    extra_b = extra_b or []

    def av(rec, c):
        return rec.get(c, "") if (rec and c) else ""

    table = []
    for r in rows:
        ra, rb = r["ra"], r["rb"]
        row = {
            "Estado": r["status"],
            "Confianza %": r["conf"],
            "Margen %": r["margin"],
            "Método": r["method"],
            "Nombre (A)": ra["_name"],
            "Nombre (B)": rb["_name"] if rb else "— sin pareja —",
            "Sim. nombre %": r["name_sim"],
            "ID/SSN (A)": av(ra, m.get("sa")),
            "Employee ID (B)": av(rb, m.get("sb")),
            "Tarifa (A)": av(ra, m.get("ra")),
            "Tarifa (B)": av(rb, m.get("rb")),
            "¿Tarifa?": "SI" if r["rate_match"] else "NO",
            "Monto (A)": av(ra, m.get("aa")),
            "Monto (B)": av(rb, m.get("ab")),
            "¿Monto?": "SI" if r["amt_match"] else "NO",
            "Semana (A)": av(ra, m.get("pa")),
            "Semana (B)": av(rb, m.get("pb")),
            "¿Semana?": "SI" if r["per_match"] else "NO",
            "¿Ambiguo?": "SI" if r["ambiguous"] else "NO",
            "¿Homónimo resuelto?": "SI" if (r["homonym"] and not r["ambiguous"]) else "NO",
        }
        for c in extra_b:
            row[f"{c} (B)"] = av(rb, c)
        table.append(row)

    for rb in b_only:
        row = {
            "Estado": "REVISAR MANUAL", "Confianza %": 0, "Margen %": 100,
            "Método": "SOLO EN B", "Nombre (A)": "— sin pareja —",
            "Nombre (B)": rb["_name"], "Sim. nombre %": 0,
            "ID/SSN (A)": "", "Employee ID (B)": av(rb, m.get("sb")),
            "Tarifa (A)": "", "Tarifa (B)": av(rb, m.get("rb")), "¿Tarifa?": "NO",
            "Monto (A)": "", "Monto (B)": av(rb, m.get("ab")), "¿Monto?": "NO",
            "Semana (A)": "", "Semana (B)": av(rb, m.get("pb")), "¿Semana?": "NO",
            "¿Ambiguo?": "NO", "¿Homónimo resuelto?": "NO",
        }
        for c in extra_b:
            row[f"{c} (B)"] = av(rb, c)
        table.append(row)

    # ordena: primero lo que hay que revisar, luego automáticos, por nombre
    table.sort(key=lambda x: (x["Estado"] != "REVISAR MANUAL", str(x["Nombre (A)"])))
    return table


# secciones para la banda de encabezado (titulo -> nº de columnas que abarca)
def _sections(extra_b):
    return [("ESTADO DEL MATCH", 4), ("IDENTIDAD", 3),
            ("VALIDACIONES  ·  A vs B", 13), ("DATOS DE B", len(extra_b))]


def write_xlsx(path, readable, full_rows):
    """Escribe un Excel formateado: hoja 'Vinculado' (legible, columnas
    pareadas, colores) + hoja 'Detalle completo' (todas las columnas)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1E293B"; BLUE = "3B82F6"; SLATE = "475569"
    GREEN_BG = "DCFCE7"; GREEN_TX = "166534"
    AMBER_BG = "FEF3C7"; AMBER_TX = "92400E"
    RED_BG = "FEE2E2"; RED_TX = "991B1B"; GREY = "F1F5F9"
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(readable[0].keys()) if readable else []
    extra_n = max(0, len(headers) - 20)
    sections = _sections(["x"] * extra_n)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vinculado"

    # fila 1: banda de secciones (merge)
    col = 1
    for title, span in sections:
        if span <= 0:
            continue
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        c = ws.cell(row=1, column=col, value=title)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=SLATE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += span

    # fila 2: encabezados
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # filas de datos
    for i, row in enumerate(readable, start=3):
        manual = row["Estado"] == "REVISAR MANUAL"
        for j, h in enumerate(headers, 1):
            val = row.get(h, "")
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = border
            c.alignment = Alignment(horizontal="center" if h not in
                                    ("Nombre (A)", "Nombre (B)") else "left",
                                    vertical="center")
            if i % 2 == 0 and not manual:
                c.fill = PatternFill("solid", fgColor=GREY)
            # color del estado
            if h == "Estado":
                if val == "AUTOMATICO":
                    c.fill = PatternFill("solid", fgColor=GREEN_BG)
                    c.font = Font(name="Arial", bold=True, size=10, color=GREEN_TX)
                else:
                    c.fill = PatternFill("solid", fgColor=AMBER_BG)
                    c.font = Font(name="Arial", bold=True, size=10, color=AMBER_TX)
            # color de las columnas ¿coincide?
            if h in ("¿Tarifa?", "¿Monto?", "¿Semana?"):
                if val == "SI":
                    c.fill = PatternFill("solid", fgColor=GREEN_BG); c.font = Font(name="Arial", size=10, color=GREEN_TX)
                else:
                    c.fill = PatternFill("solid", fgColor=RED_BG); c.font = Font(name="Arial", size=10, color=RED_TX)
            if h == "¿Ambiguo?" and val == "SI":
                c.fill = PatternFill("solid", fgColor=RED_BG); c.font = Font(name="Arial", bold=True, size=10, color=RED_TX)
            if h == "¿Homónimo resuelto?" and val == "SI":
                c.fill = PatternFill("solid", fgColor=AMBER_BG); c.font = Font(name="Arial", size=10, color=AMBER_TX)

    # anchos
    widths = {"Estado": 16, "Confianza %": 11, "Margen %": 10, "Método": 22,
              "Nombre (A)": 26, "Nombre (B)": 26, "Sim. nombre %": 11,
              "ID/SSN (A)": 14, "Employee ID (B)": 14,
              "Tarifa (A)": 10, "Tarifa (B)": 10, "¿Tarifa?": 9,
              "Monto (A)": 11, "Monto (B)": 11, "¿Monto?": 9,
              "Semana (A)": 13, "Semana (B)": 13, "¿Semana?": 9,
              "¿Ambiguo?": 10, "¿Homónimo resuelto?": 13}
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 18)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # hoja 2: detalle completo
    if full_rows:
        ws2 = wb.create_sheet("Detalle completo")
        fh = list(full_rows[0].keys())
        for j, h in enumerate(fh, 1):
            c = ws2.cell(row=1, column=j, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=NAVY)
        for i, row in enumerate(full_rows, start=2):
            for j, h in enumerate(fh, 1):
                ws2.cell(row=i, column=j, value=row.get(h, "")).font = Font(name="Arial", size=9)
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(fh))}1"

    wb.save(path)
    return path


def autodetect(columns, hints):
    low = {c: str(c).strip().lower() for c in columns}
    for h in hints:
        for c, cl in low.items():
            if cl == h:
                return c
    for h in hints:
        for c, cl in low.items():
            if h in cl:
                return c
    return None


# ===========================================================================
#  INTERFAZ GRAFICA  (Tkinter)
# ===========================================================================

def run_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    BG, PANEL, CARD = "#0f172a", "#1e293b", "#273449"
    ACCENT, GREEN, AMBER, RED = "#3b82f6", "#22c55e", "#f59e0b", "#ef4444"
    TXT, MUTE = "#e2e8f0", "#94a3b8"

    app = tk.Tk()
    app.title("Vinculador de Registros · Primary Staffing")
    app.geometry("1180x780")
    app.configure(bg=BG)
    app.minsize(980, 660)

    state = {
        "path_a": tk.StringVar(), "path_b": tk.StringVar(),
        "df_a": None, "df_b": None,
        "name_a": tk.StringVar(), "amt_a": tk.StringVar(), "sid_a": tk.StringVar(),
        "rate_a": tk.StringVar(), "per_a": tk.StringVar(),
        "name_b": tk.StringVar(), "amt_b": tk.StringVar(), "sid_b": tk.StringVar(),
        "rate_b": tk.StringVar(), "per_b": tk.StringVar(),
        "threshold": tk.IntVar(value=90),
        "out_table": None,
    }
    q = queue.Queue()

    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(".", background=PANEL, foreground=TXT, fieldbackground=PANEL,
                    bordercolor=CARD, font=("Segoe UI", 10))
    style.configure("TFrame", background=PANEL)
    style.configure("TLabel", background=PANEL, foreground=TXT)
    style.configure("Mute.TLabel", background=PANEL, foreground=MUTE)
    style.configure("Title.TLabel", background=BG, foreground=TXT, font=("Segoe UI Semibold", 18))
    style.configure("Sub.TLabel", background=BG, foreground=MUTE, font=("Segoe UI", 10))
    style.configure("Accent.TButton", background=ACCENT, foreground="white",
                    font=("Segoe UI Semibold", 11), borderwidth=0, padding=10)
    style.map("Accent.TButton", background=[("active", "#2563eb"), ("disabled", "#334155")])
    style.configure("TButton", background=CARD, foreground=TXT, borderwidth=0, padding=7)
    style.map("TButton", background=[("active", "#33415a")])
    style.configure("TCombobox", fieldbackground=PANEL, background=CARD,
                    foreground=TXT, arrowcolor=TXT)
    style.configure("Treeview", background=PANEL, fieldbackground=PANEL,
                    foreground=TXT, rowheight=26, borderwidth=0)
    style.configure("Treeview.Heading", background=CARD, foreground=TXT,
                    font=("Segoe UI Semibold", 10), borderwidth=0)
    style.map("Treeview", background=[("selected", ACCENT)])
    style.configure("Horizontal.TProgressbar", background=ACCENT, troughcolor=CARD, borderwidth=0)

    head = tk.Frame(app, bg=BG)
    head.pack(fill="x", padx=22, pady=(18, 4))
    ttk.Label(head, text="Vinculador de Registros", style="Title.TLabel").pack(anchor="w")
    ttk.Label(head, text="Clave principal: NOMBRE  ·  confirma con monto e ID secundario  ·  los dudosos van a revisión manual",
              style="Sub.TLabel").pack(anchor="w")

    body = tk.Frame(app, bg=BG)
    body.pack(fill="both", expand=True, padx=22, pady=12)
    body.columnconfigure(0, weight=0, minsize=390)
    body.columnconfigure(1, weight=1)
    body.rowconfigure(0, weight=1)

    left = tk.Frame(body, bg=PANEL)
    left.grid(row=0, column=0, sticky="nsew", padx=(0, 14))

    def section(parent, title):
        f = tk.Frame(parent, bg=PANEL)
        f.pack(fill="x", padx=16, pady=(14, 0))
        ttk.Label(f, text=title, style="Mute.TLabel",
                  font=("Segoe UI Semibold", 10)).pack(anchor="w")
        return f

    def file_picker(parent, label, path_var, on_load):
        sec = section(parent, label)
        row = tk.Frame(sec, bg=PANEL); row.pack(fill="x", pady=6)
        ent = tk.Entry(row, textvariable=path_var, bg=CARD, fg=TXT,
                       insertbackground=TXT, relief="flat")
        ent.pack(side="left", fill="x", expand=True, ipady=6, padx=(0, 8))

        def browse():
            p = filedialog.askopenfilename(
                title=f"Selecciona {label}",
                filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv *.tsv"), ("Todos", "*.*")])
            if p:
                path_var.set(p); on_load(p)
        ttk.Button(row, text="Examinar…", command=browse).pack(side="left")

    def col_mapper(parent, name_v, amt_v, sid_v, rate_v, per_v):
        f = tk.Frame(parent, bg=PANEL); f.pack(fill="x", padx=16, pady=(2, 4))
        combos = {}
        rows_def = [("Nombre (clave principal)", name_v),
                    ("Monto confirmación (opcional)", amt_v),
                    ("Tarifa/Rate confirmación (opc.)", rate_v),
                    ("Semana/Periodo (opcional)", per_v),
                    ("ID secundario SSN/EmpID (opc.)", sid_v)]
        for i, (lbl, var) in enumerate(rows_def):
            ttk.Label(f, text=lbl, style="Mute.TLabel").grid(row=i, column=0, sticky="w",
                                                             pady=3, padx=(0, 8))
            cb = ttk.Combobox(f, textvariable=var, state="readonly", width=22)
            cb.grid(row=i, column=1, sticky="ew", pady=3)
            combos[lbl] = cb
        f.columnconfigure(1, weight=1)
        return combos

    def load_file(which, path):
        try:
            df = load_table(path)
        except Exception as e:
            messagebox.showerror("Error al leer archivo", str(e)); return
        cols = list(df.columns)
        state[f"df_{which}"] = df
        nm = autodetect(cols, HINTS_NAME)
        amt = autodetect(cols, HINTS_AMOUNT)
        sid = autodetect(cols, HINTS_SID)
        rate = autodetect(cols, HINTS_RATE)
        per = autodetect(cols, HINTS_PERIOD)
        combos = combos_a if which == "a" else combos_b
        for cb in combos.values():
            cb["values"] = [""] + cols
        state[f"name_{which}"].set(nm or (cols[0] if cols else ""))
        state[f"amt_{which}"].set(amt or "")
        state[f"sid_{which}"].set(sid or "")
        state[f"rate_{which}"].set(rate or "")
        state[f"per_{which}"].set(per or "")
        log(f"✓ Archivo {which.upper()}: {os.path.basename(path)} ({len(df)} filas, {len(cols)} columnas)")
        log(f"   Nombre: {nm} | Monto: {amt} | Tarifa: {rate} | Semana: {per} | ID: {sid}")

    file_picker(left, "ARCHIVO A", state["path_a"], lambda p: load_file("a", p))
    combos_a = col_mapper(left, state["name_a"], state["amt_a"], state["sid_a"],
                          state["rate_a"], state["per_a"])
    file_picker(left, "ARCHIVO B", state["path_b"], lambda p: load_file("b", p))
    combos_b = col_mapper(left, state["name_b"], state["amt_b"], state["sid_b"],
                          state["rate_b"], state["per_b"])

    th_sec = section(left, "UMBRAL DE CONFIANZA (auto vs. manual)")
    th_row = tk.Frame(th_sec, bg=PANEL); th_row.pack(fill="x", pady=6)
    th_lbl = ttk.Label(th_row, text="90%", font=("Segoe UI Semibold", 12))
    th_lbl.pack(side="right")
    ttk.Scale(th_row, from_=70, to=100, orient="horizontal", variable=state["threshold"],
              command=lambda v: th_lbl.config(text=f"{int(float(v))}%")).pack(
        side="left", fill="x", expand=True, padx=(0, 10))
    ttk.Label(left, text="Arriba del umbral → AUTOMÁTICO.  Debajo → REVISAR MANUAL.",
              style="Mute.TLabel", font=("Segoe UI", 9)).pack(anchor="w", padx=16)

    btn_run = ttk.Button(left, text="▶  Procesar y vincular", style="Accent.TButton",
                         command=lambda: start_process())
    btn_run.pack(fill="x", padx=16, pady=(18, 4))
    prog = ttk.Progressbar(left, style="Horizontal.TProgressbar", maximum=100)
    prog.pack(fill="x", padx=16, pady=(6, 2))
    prog_lbl = ttk.Label(left, text="Listo.", style="Mute.TLabel", font=("Segoe UI", 9))
    prog_lbl.pack(anchor="w", padx=16)

    right = tk.Frame(body, bg=PANEL)
    right.grid(row=0, column=1, sticky="nsew")
    right.rowconfigure(2, weight=3); right.rowconfigure(5, weight=1)
    right.columnconfigure(0, weight=1)

    cards = tk.Frame(right, bg=PANEL)
    cards.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 6))
    card_vals = {}

    def make_card(key, title, color):
        c = tk.Frame(cards, bg=CARD, highlightbackground=color, highlightthickness=2)
        c.pack(side="left", expand=True, fill="x", padx=4)
        v = tk.Label(c, text="—", bg=CARD, fg=color, font=("Segoe UI Semibold", 22))
        v.pack(anchor="w", padx=12, pady=(8, 0))
        tk.Label(c, text=title, bg=CARD, fg=MUTE, font=("Segoe UI", 9)).pack(
            anchor="w", padx=12, pady=(0, 8))
        card_vals[key] = v

    make_card("auto", "Automáticos", GREEN)
    make_card("manual", "Revisar manual", AMBER)
    make_card("pct", "% Auto-match", ACCENT)
    make_card("total", "Total A", TXT)

    ttk.Label(right, text="RESULTADOS", style="Mute.TLabel",
              font=("Segoe UI Semibold", 10)).grid(row=1, column=0, sticky="w",
                                                   padx=16, pady=(8, 2))
    tbl = tk.Frame(right, bg=PANEL); tbl.grid(row=2, column=0, sticky="nsew", padx=16)
    cols = ("nombre", "metodo", "conf", "margen", "sim", "estado")
    tree = ttk.Treeview(tbl, columns=cols, show="headings")
    for c, txt, w in [("nombre", "Nombre (A)", 230), ("metodo", "Método", 130),
                      ("conf", "Confianza %", 95), ("margen", "Margen err %", 100),
                      ("sim", "Sim. nombre %", 100), ("estado", "Estado", 140)]:
        tree.heading(c, text=txt); tree.column(c, width=w, anchor="w")
    vsb = ttk.Scrollbar(tbl, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")
    tree.tag_configure("auto", foreground=GREEN)
    tree.tag_configure("manual", foreground=AMBER)
    tree.tag_configure("none", foreground=RED)

    exp = tk.Frame(right, bg=PANEL); exp.grid(row=3, column=0, sticky="ew", padx=16, pady=(8, 4))
    btn_xlsx = ttk.Button(exp, text="⬇  Exportar Excel organizado",
                          style="Accent.TButton",
                          command=lambda: export_xlsx(), state="disabled")
    btn_xlsx.pack(side="left", padx=(0, 8))
    btn_csv = ttk.Button(exp, text="⬇  CSV consolidado",
                         command=lambda: export_csv(False), state="disabled")
    btn_csv.pack(side="left", padx=(0, 8))
    btn_rev = ttk.Button(exp, text="⬇  Solo 'Revisar manual'",
                         command=lambda: export_csv(True), state="disabled")
    btn_rev.pack(side="left")

    ttk.Label(right, text="REGISTRO DE ACTIVIDAD", style="Mute.TLabel",
              font=("Segoe UI Semibold", 10)).grid(row=4, column=0, sticky="nw",
                                                   padx=16, pady=(8, 2))
    log_box = tk.Text(right, height=7, bg="#0b1220", fg="#9fb3c8", insertbackground=TXT,
                      relief="flat", font=("Consolas", 9), wrap="word")
    log_box.grid(row=5, column=0, sticky="nsew", padx=16, pady=(0, 14))

    def log(msg):
        log_box.insert("end", f"[{dt.datetime.now():%H:%M:%S}] {msg}\n"); log_box.see("end")

    log("Carga el Archivo A y el B. La clave principal es el NOMBRE. "
        "El monto y el ID secundario solo confirman.")
    if not HAVE_RF:
        log("⚠ rapidfuzz no instalado: comparación básica. Mejor: pip install rapidfuzz")

    def start_process():
        if state["df_a"] is None or state["df_b"] is None:
            messagebox.showwarning("Faltan archivos", "Carga el Archivo A y el Archivo B."); return
        if not state["name_a"].get() or not state["name_b"].get():
            messagebox.showwarning("Falta nombre", "Selecciona la columna de Nombre en ambos archivos."); return
        btn_run.config(state="disabled"); btn_csv.config(state="disabled"); btn_rev.config(state="disabled")
        for i in tree.get_children():
            tree.delete(i)
        prog["value"] = 0
        log("──────── Iniciando vinculación ────────")

        def worker():
            try:
                df_a, df_b = state["df_a"], state["df_b"]
                recs_a = parse_blocks(df_a, state["name_a"].get(),
                                      state["amt_a"].get() or None, state["sid_a"].get() or None,
                                      state["rate_a"].get() or None, state["per_a"].get() or None)
                q.put(("log", f"   {len(recs_a)} personas en A."))
                recs_b = parse_blocks(df_b, state["name_b"].get(),
                                      state["amt_b"].get() or None, state["sid_b"].get() or None,
                                      state["rate_b"].get() or None, state["per_b"].get() or None)
                q.put(("log", f"   {len(recs_b)} personas en B."))
                q.put(("log", "Buscando coincidencias por NOMBRE + monto + tarifa + semana…"))
                rows, b_only, shared = link(recs_a, recs_b, state["threshold"].get(),
                                            lambda d, t, nm: q.put(("prog", (d, t, nm))))
                if shared:
                    q.put(("log", "✓ Ambos archivos comparten un ID secundario: confirma al 100%."))
                else:
                    q.put(("log", "ℹ Sin ID compartido: se confía en nombre + monto + tarifa."))
                out = build_output(rows, b_only, list(df_a.columns), list(df_b.columns))
                m = {"na": state["name_a"].get(), "aa": state["amt_a"].get() or None,
                     "ra": state["rate_a"].get() or None, "pa": state["per_a"].get() or None,
                     "sa": state["sid_a"].get() or None,
                     "nb": state["name_b"].get(), "ab": state["amt_b"].get() or None,
                     "rb": state["rate_b"].get() or None, "pb": state["per_b"].get() or None,
                     "sb": state["sid_b"].get() or None}
                used = set(filter(None, [m["nb"], m["ab"], m["rb"], m["pb"], m["sb"]]))
                extra_b = [c for c in df_b.columns if c not in used][:6]
                readable = build_readable(rows, b_only, m, extra_b)
                q.put(("done", (rows, b_only, out, readable)))
            except Exception as e:
                q.put(("error", str(e)))

        threading.Thread(target=worker, daemon=True).start()
        app.after(60, poll_queue)

    def poll_queue():
        try:
            while True:
                kind, payload = q.get_nowait()
                if kind == "log":
                    log(payload)
                elif kind == "prog":
                    d, t, nm = payload
                    prog["value"] = 100 * d / max(t, 1)
                    prog_lbl.config(text=f"Procesando {d}/{t} · {str(nm)[:28]}")
                elif kind == "error":
                    log(f"✗ ERROR: {payload}")
                    messagebox.showerror("Error", payload)
                    btn_run.config(state="normal"); return
                elif kind == "done":
                    finish(*payload); return
        except queue.Empty:
            pass
        app.after(60, poll_queue)

    def finish(rows, b_only, out, readable):
        state["out_table"] = out
        state["readable"] = readable
        prog["value"] = 100; prog_lbl.config(text="Completado.")
        total = len(rows)
        auto = sum(1 for r in rows if r["status"] == "AUTOMATICO")
        manual = total - auto
        pct = 100 * auto / total if total else 0
        card_vals["auto"].config(text=str(auto))
        card_vals["manual"].config(text=str(manual + len(b_only)))
        card_vals["pct"].config(text=f"{pct:.0f}%")
        card_vals["total"].config(text=str(total))

        for r in rows:
            tag = "auto" if r["status"] == "AUTOMATICO" else (
                "none" if r["method"] == "SIN MATCH" else "manual")
            tree.insert("", "end", tags=(tag,), values=(
                r["ra"]["_name"], r["method"], f'{r["conf"]:.1f}',
                f'{r["margin"]:.1f}', f'{r["name_sim"]:.0f}', r["status"]))
        for rb in b_only:
            tree.insert("", "end", tags=("none",), values=(
                rb["_name"] + "  (solo en B)", "SIN MATCH", "0.0", "100.0", "0", "REVISAR MANUAL"))

        n_id = sum(1 for r in rows if r["id_match"])
        n_amt = sum(1 for r in rows if r["amt_match"])
        n_rate = sum(1 for r in rows if r["rate_match"])
        n_per = sum(1 for r in rows if r["per_match"])
        n_amb = sum(1 for r in rows if r["ambiguous"])
        log("──────── Resultado ────────")
        log(f"Confirmaciones → ID: {n_id} · monto: {n_amt} · tarifa: {n_rate} · semana: {n_per}")
        log(f"Nombres ambiguos (a revisión): {n_amb}")
        log(f"AUTOMÁTICOS: {auto} ({pct:.1f}%) · MANUAL: {manual + len(b_only)} ({100 - pct:.1f}% + {len(b_only)} solo-en-B)")
        log("Revisa los ámbar/rojo y exporta el Excel organizado.")
        btn_run.config(state="normal"); btn_xlsx.config(state="normal")
        btn_csv.config(state="normal"); btn_rev.config(state="normal")

    def export_xlsx():
        if not state.get("readable"):
            return
        default = f"Vinculado_organizado_{dt.datetime.now():%Y%m%d_%H%M}.xlsx"
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default,
                                            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            write_xlsx(path, state["readable"], state["out_table"])
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e)); return
        log(f"✓ Excel organizado: {os.path.basename(path)} ({len(state['readable'])} filas)")
        messagebox.showinfo("Exportado",
                            f"Guardado:\n{path}\n\nHoja 'Vinculado' (legible) + "
                            f"'Detalle completo'.")

    def export_csv(only_review):
        if not state["out_table"]:
            return
        default = ("revisar_manual" if only_review else "vinculado_consolidado")
        default += f"_{dt.datetime.now():%Y%m%d_%H%M}.csv"
        path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=default,
                                            filetypes=[("CSV", "*.csv")])
        if not path:
            return
        rows = state["readable"] if not only_review else \
            [r for r in state["readable"] if r["Estado"] == "REVISAR MANUAL"]
        if not rows:
            messagebox.showinfo("Sin filas", "No hay filas que exportar."); return
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader(); w.writerows(rows)
        log(f"✓ Exportado: {os.path.basename(path)} ({len(rows)} filas)")
        messagebox.showinfo("Exportado", f"Guardado:\n{path}\n\n{len(rows)} filas.")

    app.mainloop()


if __name__ == "__main__":
    run_gui()